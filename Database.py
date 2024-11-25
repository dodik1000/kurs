import customtkinter as ctk
from openpyxl import load_workbook
from CTkMessagebox import CTkMessagebox

""" Класс для работы с базой данных Excel """


class Database:
    def __init__(self, parent, fx, a, b, n):
        """ Инициализация базы данных """
        self.fn = 'materials/database.xlsx'
        self.wb = load_workbook(self.fn)
        self.ws = self.wb['Sheet1']
        self.new_window = None
        self.parent = parent
        self.fx = fx
        self.a = a
        self.b = b
        self.n = n
        self.new_win_counter = 0

    def load_history_data(self):
        """ Чтение данных из Excel файла """
        for i, row in enumerate(self.ws.iter_rows(min_row=2, max_row=11)):
            if i < len(self.history_frames):
                fx = row[0].value
                a = row[1].value
                b = row[2].value
                n = row[3].value

                # Обновление текста лейбла
                history_label, _ = self.history_frames[i]
                history_label.configure(text=f"f(x) = {fx}\n"
                                             f"a = {a}\nb = {b}\nn = {n}\n")

        self.wb.close()

    def history_window(self):
        """ Создание окна истории """
        self.new_window = ctk.CTk()
        self.new_window.title("История вычислений")
        self.new_window.geometry("300x600")
        self.new_window.resizable(width=False, height=False)
        self.new_window.deiconify()
        self.new_window.protocol("WM_DELETE_WINDOW", self.close_history_window)

        self.history_frames = []

        # Создание фрейма с прокруткой
        fg_color = "#202020" if ctk.get_appearance_mode() == "Dark" else "#f3f3f3"

        self.scrollable_frame = ctk.CTkScrollableFrame(self.new_window, width=280,
                                                       fg_color=fg_color, height=540)
        self.scrollable_frame.grid(row=0, column=0)

        # Создание 10 фреймов
        for i in range(10):
            self.history_frame = ctk.CTkFrame(self.scrollable_frame, corner_radius=10,
                                              height=100, width=270)
            self.history_frame.grid(row=i, column=0, padx=5, pady=5)
            self.history_frame.grid_propagate(False)

            # Установка конфигурации столбцов и строк внутри фрейма
            self.history_frame.grid_columnconfigure(0, weight=1)
            self.history_frame.grid_columnconfigure(1, weight=0)
            self.history_frame.grid_rowconfigure(0, weight=1)

            self.history_label = ctk.CTkLabel(self.history_frame, text="",
                                              font=("Arial Black", 12),
                                              wraplength=220, justify="left")
            self.history_label.grid(row=0, column=0, padx=10, pady=10, sticky="w")

            self.history_button = ctk.CTkButton(self.history_frame, text="Выбрать",
                                                font=("Arial Black", 12),
                                                command=lambda idx=i:
                                                self.fill_input_fields(idx),
                                                fg_color="#7d748e",
                                                hover_color="#545164", width=80)
            self.history_button.grid(row=0, column=1, padx=10, pady=10, sticky="e")

            self.history_frames.append((self.history_label, self.history_button))

        self.clear_history_button = ctk.CTkButton(self.new_window,
                                                  text="Очистить историю",
                                                  command=self.clear_history,
                                                  fg_color="#7d748e",
                                                  hover_color="#545164", width=150)
        self.clear_history_button.grid(row=1, column=0,
                                       padx=10, pady=10, sticky="nsew")

        # Загрузка данных истории
        self.load_history_data()

        return self.new_window

    def clear_history(self):
        """ Очищение ячейки с данными истории """
        for row in range(2, self.ws.max_row + 1):
            # Очищаем столбцы A-D
            for col in range(1, 5):
                self.ws.cell(row=row, column=col).value = None

        # Сброс счетчика
        self.ws['F2'].value = 2

        self.wb.save(self.fn)
        self.wb.close()

        # Обновление отображения истории в окне
        self.load_history_data()

        # Сброс текста лейбла в окне истории
        for history_label, _ in self.history_frames:
            history_label.configure(text="")

    def fill_input_fields(self, button_index):
        """ Заполнение полей в calculate_frame данными из истории """
        if button_index < len(self.history_frames):
            row = list(self.ws.iter_rows(min_row=button_index + 2,
                                         max_row=button_index + 2))
            if row:
                r = row[0]
                self.fx = r[0].value
                self.a = r[1].value
                self.b = r[2].value
                self.n = r[3].value

                # Проверка на наличие данных в ячейках
                if (self.fx is None or self.a is None or
                        self.b is None or self.n is None):
                    CTkMessagebox(title="Ошибка", width=300,
                                  message="В выбранной строке отсутствуют данные!",
                                  height=200, icon="cancel")
                    return

                # Заполнение полей в calculate_frame
                self.parent.entry_func.delete(0, 'end')
                self.parent.entry_func.insert(0, self.fx)
                self.parent.entry_a.delete(0, 'end')
                self.parent.entry_a.insert(0, self.a)
                self.parent.entry_b.delete(0, 'end')
                self.parent.entry_b.insert(0, self.b)
                self.parent.entry_n.delete(0, 'end')
                self.parent.entry_n.insert(0, self.n)

            self.wb.close()

    def update_history(self):
        """ Обновление истории в базе данных """
        row_counter_cell = self.ws['F2']
        self.row_counter = row_counter_cell.value if row_counter_cell.value else 2

        # Установка значения в новую строку
        self.ws[f'A{self.row_counter}'] = self.fx
        self.ws[f'B{self.row_counter}'] = self.a
        self.ws[f'C{self.row_counter}'] = self.b
        self.ws[f'D{self.row_counter}'] = self.n

        # Увеличение счетчика
        self.row_counter += 1

        # Проверка счетчика
        if self.row_counter > 11:
            self.row_counter = 2

        self.ws['F2'] = self.row_counter

        self.wb.save(self.fn)
        self.wb.close()

    def close_history_window(self):
        """ Закрытие окна истории """
        self.new_window.destroy()
        self.new_window = None
        self.parent.new_window_deleter = None
